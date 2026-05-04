"""
Batch-load test cases from test_cases.tsv into 'Assignment 1 - Test cases.xlsx'.

The TSV file is tab-separated with this header (4 columns):
    TC ID    Input length type    Input    Expected output

This is the easiest way to manage 50+ test cases — edit them in any text
editor (or Excel exported as TSV) and run this once to populate the
assignment template.

Usage:
    python load_test_cases.py             # loads test_cases.tsv
    python load_test_cases.py my.tsv      # loads a different file
    python load_test_cases.py --reset     # clear test data, keep template
    python load_test_cases.py --run       # load TSV, then run Playwright tests
"""
import sys
import re
import subprocess
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Side

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
except Exception:
    pass

FIELD_SPLIT = re.compile(r"\t+| {2,}")

ROOT = Path(__file__).resolve().parent
EXCEL = ROOT / "Assignment 1 - Test cases.xlsx"
DEFAULT_TSV = ROOT / "test_cases.tsv"
SHEET = " Test cases"
HEADER_ROW = 1
BLOCK_HEIGHT = 4
DATA_COLS = [1, 2, 3, 4, 5, 6]  # TC ID, Length, Input, Expected, Actual, Status


def _ensure_block_merges(ws, top_row):
    bottom = top_row + BLOCK_HEIGHT - 1
    existing = {str(r) for r in ws.merged_cells.ranges}
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in DATA_COLS:
        col_letter = ws.cell(row=top_row, column=col).column_letter
        rng = f"{col_letter}{top_row}:{col_letter}{bottom}"
        if rng not in existing:
            ws.merge_cells(rng)
        ws.cell(row=top_row, column=col).alignment = center
        for r in range(top_row, bottom + 1):
            ws.cell(row=r, column=col).border = border


def _clear_test_data(ws):
    from openpyxl.cell.cell import MergedCell
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        for c in DATA_COLS:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _check_writable(path: Path):
    try:
        with open(path, "ab"):
            pass
    except PermissionError:
        print(f"ERROR: '{path.name}' is open in Excel. Close it first, then re-run.")
        sys.exit(1)


def _last_filled_top_row(ws) -> int:
    """Return the top row of the last filled test-case block (or HEADER_ROW if none)."""
    last = HEADER_ROW
    r = HEADER_ROW + 1
    while r <= ws.max_row:
        cell = ws.cell(row=r, column=1)
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            r += 1
            continue
        if cell.value is not None and str(cell.value).strip():
            last = r
            r += BLOCK_HEIGHT
        else:
            r += 1
    return last


def load(tsv_path: Path, append: bool = False, only_ids: set | None = None):
    _check_writable(EXCEL)
    if not tsv_path.exists():
        print(f"ERROR: TSV file not found: {tsv_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    if append:
        last_top = _last_filled_top_row(ws)
        top = last_top + BLOCK_HEIGHT if last_top > HEADER_ROW else HEADER_ROW + 1
        print(f"Append mode: starting at row {top}")
    else:
        _clear_test_data(ws)
        top = HEADER_ROW + 1

    with open(tsv_path, encoding="utf-8") as f:
        lines = f.read().splitlines()

    if not lines:
        print("ERROR: TSV file is empty.")
        sys.exit(1)

    header = FIELD_SPLIT.split(lines[0].strip(), maxsplit=3)
    if len(header) < 4:
        print("ERROR: TSV header missing. Expected 4 columns separated by tabs (or 2+ spaces).")
        sys.exit(1)

    added = 0
    for line_no, raw in enumerate(lines[1:], start=2):
        if not raw.strip():
            continue
        parts = FIELD_SPLIT.split(raw.rstrip(), maxsplit=3)
        if len(parts) < 4:
            print(f"  line {line_no}: skipped (need 4 fields, got {len(parts)} — separate with TAB or 2+ spaces)")
            continue
        tc_id, length, singlish, expected = parts[0].strip(), parts[1].strip(), parts[2].strip(), parts[3].strip()
        if length not in {"S", "M", "L"}:
            print(f"  line {line_no} ({tc_id}): bad length '{length}' — must be S/M/L. Skipping.")
            continue
        if only_ids is not None and tc_id not in only_ids:
            continue
        _ensure_block_merges(ws, top)
        ws.cell(row=top, column=1).value = tc_id
        ws.cell(row=top, column=2).value = length
        ws.cell(row=top, column=3).value = singlish
        ws.cell(row=top, column=4).value = expected
        try:
            print(f"  row {top}: {tc_id}  [{length}]  {singlish[:40]}...")
        except UnicodeEncodeError:
            safe = singlish[:40].encode("ascii", "backslashreplace").decode("ascii")
            print(f"  row {top}: {tc_id}  [{length}]  {safe}...")
        added += 1
        top += BLOCK_HEIGHT

    wb.save(EXCEL)
    print(f"\nLoaded {added} test cases into '{EXCEL.name}'.")


def reset():
    _check_writable(EXCEL)
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    _clear_test_data(ws)
    wb.save(EXCEL)
    print(f"Cleared all test data from '{EXCEL.name}'. Template preserved.")


def run_automation(extra_args: list[str]):
    script = ROOT / "test_automation.py"
    cmd = [sys.executable, str(script), *extra_args]
    print(f"\nRunning: {' '.join(cmd)}\n")
    subprocess.run(cmd, check=False)


def _parse_only(args: list[str]) -> tuple[set | None, list[str]]:
    """Extract --only=ID1,ID2,... or --only ID1,ID2,..."""
    only: set | None = None
    out: list[str] = []
    i = 0
    while i < len(args):
        a = args[i]
        if a.startswith("--only="):
            only = {x.strip() for x in a.split("=", 1)[1].split(",") if x.strip()}
        elif a == "--only" and i + 1 < len(args):
            only = {x.strip() for x in args[i + 1].split(",") if x.strip()}
            i += 1
        else:
            out.append(a)
        i += 1
    return only, out


if __name__ == "__main__":
    args = sys.argv[1:]
    if "--reset" in args:
        reset()
    else:
        run = "--run" in args
        append = "--append" in args
        if run:
            args = [a for a in args if a != "--run"]
        if append:
            args = [a for a in args if a != "--append"]
        only_ids, args = _parse_only(args)
        tsv_arg = next((a for a in args if not a.startswith("--")), None)
        tsv = Path(tsv_arg) if tsv_arg else DEFAULT_TSV
        if not tsv.is_absolute():
            tsv = ROOT / tsv
        load(tsv, append=append, only_ids=only_ids)
        if run:
            forwarded = [a for a in args if a.startswith("--")]
            if only_ids:
                forwarded = forwarded + ["--only", ",".join(sorted(only_ids))]
            run_automation(forwarded or ["--keep-open"])
